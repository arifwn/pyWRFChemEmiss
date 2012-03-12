
"""
unittest untuk core module
"""

import unittest
import domain_data
import namelist_reader
import data_downscaler
from openpyxl.reader.excel import load_workbook

class CoreTest(unittest.TestCase):
    def setUp(self):
        path = 'res/namelist.wps'
        self.wpsreader = namelist_reader.WPSNamelistReader(path)
        tdm = domain_data.TopDomain()
        print self.wpsreader.parent_id
        tdm.lat = self.wpsreader.ref_lat
        tdm.lon = self.wpsreader.ref_lon
        tdm.w   = self.wpsreader.e_we[0] - 1
        tdm.h   = self.wpsreader.e_sn[0] - 1
        tdm.dx  = self.wpsreader.dom_dx
        tdm.dy  = self.wpsreader.dom_dy
        tdm.process_boundary()
        
        cdm = domain_data.SubDomain(tdm, self.wpsreader.i_parent_start[1], 
                                    self.wpsreader.j_parent_start[1], self.wpsreader.parent_grid_ratio[1])
        cdm.w   = self.wpsreader.e_we[1] - 1
        cdm.h   = self.wpsreader.e_sn[1] - 1
        cdm.process_boundary()
        
        ccdm = domain_data.SubDomain(cdm, self.wpsreader.i_parent_start[2], 
                                    self.wpsreader.j_parent_start[2], self.wpsreader.parent_grid_ratio[2])
        ccdm.w   = self.wpsreader.e_we[2] - 1
        ccdm.h   = self.wpsreader.e_sn[2] - 1
        ccdm.process_boundary()
        
        self.tdm = tdm
        self.cdm = cdm
        self.ccdm = ccdm
        
#    def test_excel_read(self):
#        path = "res/grid_marsel-new.xlsx"
#        wb = load_workbook(path)
#        sheet = wb.get_active_sheet()
#        cells = sheet.range('D2:D9802')
#        print cells[0][0].value, type(cells[0][0].value)
#        self.assertAlmostEqual(cells[0][0].value , 11930967.6682)
#        self.assertAlmostEqual(cells[9800][0].value , 12028967.6689)
        
    def test_core(self):
        
        data_w = 99
        data_h = 99
        row_start = 2
        row_end = 9802
        conv_factor = 1000 / (365 * 24)
        range_x_str = "D{0}:D{1}".format(row_start, row_end)
        range_y_str = "E{0}:E{1}".format(row_start, row_end)
        range_lon_str = "F{0}:F{1}".format(row_start, row_end)
        range_lat_str = "G{0}:G{1}".format(row_start, row_end)
        range_conc_str = "I{0}:I{1}".format(row_start, row_end)
        print range_lat_str, range_lon_str, range_x_str, range_y_str, range_conc_str
        
        tdm  = self.tdm
        cdm  = [self.cdm, self.ccdm]
#        print tdm.get_cell_boundary_latlon(0, 0)
#        print cdm.get_cell_boundary_latlon(0, 0)
#        print ccdm.get_cell_boundary_latlon(0, 0)
        
        path = "res/grid_marsel-new.xlsx"
        
        print "Loading excel workbook..."
        wb = load_workbook(path)
        print wb.get_sheet_names()
        sheet = wb.get_active_sheet()
        
        print "Reading cell..."
        cells_x = sheet.range(range_x_str)
        cells_y = sheet.range(range_y_str)
        cells_lat = sheet.range(range_lat_str)
        cells_lon = sheet.range(range_lon_str)
        cells_conc = sheet.range(range_conc_str)
        
        data_list = []
        data_len = len(cells_x)
        
        tmp_x = 0.0
        tmp_y = 0.0
        tmp_lat = 0.0
        tmp_lon = 0.0
        tmp_conc = 0.0
        
        print "Saving list data..."
        for i in range(data_len):
            tmp_x = cells_x[i][0].value
            tmp_y = cells_y[i][0].value
            tmp_lat = cells_lat[i][0].value
            tmp_lon = cells_lon[i][0].value
            tmp_conc = cells_conc[i][0].value
            tmp_rowdata = data_downscaler.RowData(tmp_lat, tmp_lon, tmp_x, tmp_y, tmp_conc)
            data_list.append(tmp_rowdata)
        
        print "Sorting list data..."
        data_list.sort(data_downscaler.compare_rowdata_m)
        
        print "Creating emission group..."
        eg = data_downscaler.EmissGroup()
        eg.set_dimension(data_w, data_h)
        i = 0
        for y in range(data_h):
            for x in range(data_w):
                row_data = data_list[i]
                tmp = data_downscaler.EmissData()
                tmp.conc = row_data.data
                tmp.lat = row_data.lat
                tmp.lon = row_data.lon
                tmp.x = row_data.x
                tmp.y = row_data.y
                eg.append_data(tmp)
                i = i + 1
        
        eg.compute_boundary()
        
        print eg.ll_lat, eg.ll_lon
        print eg.tr_lat, eg.tr_lon
        print eg.ll_x, eg.ll_y
        print eg.tr_y, eg.tr_y
        
        print 'domain 1'
        for i in range(tdm.h):
            for j in range(tdm.w):
                ll, tr = tdm.get_cell_boundary_latlon(i, j)
                conc = eg.get_average_conc_latlon(ll, tr)
                print i, j, ':', conc
        
#        eg = data_downscaler.EmissGroup()
#        eg.set_dimension(90, 90)
#        for y in range(90):
#            for x in range(90):
#                tmp = data_downscaler.EmissData()
#                tmp.conc = y*10 + x + 5.0
#                tmp.lat = -7 + y*0.01
#                tmp.lon = 107 + x*0.01
#                tmp.x, tmp.y = utm.convert_to_utm(tmp.lat, tmp.lon)
#                eg.append_data(tmp)
        
#        for x in range(tdm.w):
#            for y in range(tdm.h):
#                print x, y, tdm.get_cell_boundary_latlon(x, y)
#        print (tdm.tr_lat, tdm.tr_lon) 
#          
#        for x in range(cdm.w):
#            for y in range(cdm.h):
#                print x, y, cdm.get_cell_boundary_latlon(x, y)
#        print (cdm.tr_lat, cdm.tr_lon)
#           
#        for x in range(ccdm.w):
#            for y in range(ccdm.h):
#                print x, y, ccdm.get_cell_boundary_latlon(x, y)
#        print (ccdm.tr_lat, ccdm.tr_lon)        
        
    
if __name__ == "__main__":
    unittest.main()
    