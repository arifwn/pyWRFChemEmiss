
from openpyxl.reader.excel import load_workbook

if __name__ == '__main__':
    path = "res/Data_Koordinat_Grid.xlsx"
    wb = load_workbook(path)
    print 'OK'
    print wb.get_sheet_names()
    sht = wb.get_active_sheet()
    cl = sht.cell('A1')
    print cl.value
    