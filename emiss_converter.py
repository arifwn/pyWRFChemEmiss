#!/usr/bin/env python
'''
Created on Mar 21, 2011

@author: Arif Widi Nugroho <arif@sainsmograf.com>
'''

import wx
import os
import cPickle
import commons
import namelist_reader
import icon
import payload
from wx.lib.wordwrap import wordwrap
from openpyxl.reader.excel import load_workbook

class PollutantData():
    def __init__(self):
        self.pollutant             = None
        self.conversion_factor     = 1.0
        self.conversion_factor_str = None
        self.worksheet             = 0
        self.row_start             = None
        self.row_end               = None
        self.emission_column       = None
        self.lat_column            = None
        self.lon_column            = None
        self.x_column              = None
        self.y_column              = None
    
    def __str__(self):
        return "{0} [{1}{2}:{1}{3} f={4}]".format(self.pollutant, self.emission_column, self.row_start, self.row_end, self.conversion_factor)

class PollutantSetting():
    def __init__(self):
        self.plt_list     = []
        self.plt_str_list = []

class MainFrame(wx.Frame):
    def __init__(self, parent=None, id=-1, title="Emission Converter"):
        wx.Frame.__init__(self, parent, id, title)
        self.panel = MainPanel(self)
        sizer = wx.BoxSizer()
        self.SetSizer(sizer)
        sizer.Add(self.panel, 1, wx.ALL | wx.EXPAND)
        self.Fit()
        app_icon = icon.get_icon_app()
        
        self.SetIcon(app_icon)

class MainPanel(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent, -1)
        
        self.file_excel    = ""
        self.file_namelist = ""
        self.workbook    = None
        self.workbook_ok = False
        self.namelist    = None
        self.namelist_ok = False
        self.size_x   = None
        self.size_y   = None
        
        self.SetBackgroundColour('White')
        sizer_main = wx.BoxSizer(wx.VERTICAL)
        self.SetSizer(sizer_main)
        
        ico = icon.get_icon_open()
        btn_choose_file = wx.BitmapButton(self, -1, ico)
        self.txt_file = wx.StaticText(self, -1, "Select *.xlsx file")
        sizer_file = wx.BoxSizer(wx.HORIZONTAL)
        sizer_file.Add(btn_choose_file, 0, wx.ALL | wx.EXPAND, 5)
        sizer_file.Add(self.txt_file, 0, wx.ALL | wx.EXPAND, 5)
        
        btn_choose_namelist = wx.BitmapButton(self, -1, ico)
        self.txt_namelist_file = wx.StaticText(self, -1, "Select namelist.wps file")
        sizer_namelist_file = wx.BoxSizer(wx.HORIZONTAL)
        sizer_namelist_file.Add(btn_choose_namelist, 0, wx.ALL | wx.EXPAND, 5)
        sizer_namelist_file.Add(self.txt_namelist_file, 0, wx.ALL | wx.EXPAND, 5)

        self.te_size_x = commons.GrayTextCtrl(self, -1, 'X size (e.g.: 100)', validator=commons.NumberTextValidator())
        self.te_size_y = commons.GrayTextCtrl(self, -1, 'Y size (e.g.: 100)', validator=commons.NumberTextValidator())
        sizer_sz = wx.BoxSizer(wx.HORIZONTAL)
        sizer_sz.Add(self.te_size_x, 1, wx.ALL | wx.EXPAND, 5)
        sizer_sz.Add(self.te_size_y, 1, wx.ALL | wx.EXPAND, 5)
        sizer_sz_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_sz_header.Add(wx.StaticText(self, -1, "Dimension "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_sz_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)

        
        self.plt_list     = []
        self.plt_str_list = []
        btn_add_pollutant = wx.Button(self, -1, "Add Pollutant")
        btn_edt_pollutant = wx.Button(self, -1, "Edit Selected")
        btn_rmv_pollutant = wx.Button(self, -1, "Remove Selected")
        sizer_btn_pol = wx.BoxSizer(wx.HORIZONTAL)
        sizer_btn_pol.Add(btn_add_pollutant, 0, wx.ALL | wx.EXPAND, 5)
        sizer_btn_pol.Add(btn_edt_pollutant, 0, wx.ALL | wx.EXPAND, 5)
        sizer_btn_pol.Add(btn_rmv_pollutant, 0, wx.ALL | wx.EXPAND, 5)
        
        self.lst_pollutant = wx.ListBox(self, -1, size=(-1, 200))
        
        btn_lst_save = wx.Button(self, -1, "Save list")
        btn_lst_open  = wx.Button(self, -1, "Open list")
        sizer_btn_lst = wx.BoxSizer(wx.HORIZONTAL)
        sizer_btn_lst.Add(btn_lst_save, 0, wx.ALL | wx.EXPAND, 5)
        sizer_btn_lst.Add(wx.BoxSizer(), 1, wx.ALL | wx.EXPAND, 5)
        sizer_btn_lst.Add(btn_lst_open, 0, wx.ALL | wx.EXPAND, 5)
        
        
        btn_about = wx.Button(self, -1, "About")
        btn_conv  = wx.Button(self, -1, "Convert!")
        sizer_btn_conv = wx.BoxSizer(wx.HORIZONTAL)
        sizer_btn_conv.Add(btn_about, 0, wx.ALL | wx.EXPAND, 5)
        sizer_btn_conv.Add(wx.BoxSizer(), 1, wx.ALL | wx.EXPAND, 5)
        sizer_btn_conv.Add(btn_conv, 0, wx.ALL | wx.EXPAND, 5)
        
        sizer_main.Add(sizer_file, 0, wx.ALL | wx.EXPAND)
        sizer_main.Add(wx.StaticLine(self, -1), 0, wx.ALL | wx.EXPAND, 5)
        sizer_main.Add(sizer_namelist_file, 0, wx.ALL | wx.EXPAND)
        sizer_main.Add(sizer_sz_header, 0, wx.ALL | wx.EXPAND, 5)
        sizer_main.Add(sizer_sz, 0, wx.ALL | wx.EXPAND, 0)
        sizer_main.Add(wx.StaticLine(self, -1), 0, wx.ALL | wx.EXPAND, 5)
        sizer_main.Add(sizer_btn_pol, 0, wx.ALL | wx.EXPAND)
        sizer_main.Add(self.lst_pollutant, 1, wx.ALL | wx.EXPAND, 5)
        sizer_main.Add(sizer_btn_lst, 0, wx.ALL | wx.EXPAND)
        sizer_main.Add(wx.StaticLine(self, -1), 0, wx.ALL | wx.EXPAND, 5)
        sizer_main.Add(sizer_btn_conv, 0, wx.ALL | wx.EXPAND)
        
        self.Bind(wx.EVT_BUTTON, self.OnButtonChoose, btn_choose_file)
        self.Bind(wx.EVT_BUTTON, self.OnButtonChooseNamelist, btn_choose_namelist)
        self.Bind(wx.EVT_BUTTON, self.OnButtonAddPollutant, btn_add_pollutant)
        self.Bind(wx.EVT_BUTTON, self.OnButtonEditPollutant, btn_edt_pollutant)
        self.Bind(wx.EVT_BUTTON, self.OnButtonRemovePollutant, btn_rmv_pollutant)
        self.Bind(wx.EVT_BUTTON, self.OnButtonAbout, btn_about)
        self.Bind(wx.EVT_BUTTON, self.OnButtonConvert, btn_conv)
        self.Bind(wx.EVT_BUTTON, self.OnButtonListOpen, btn_lst_open)
        self.Bind(wx.EVT_BUTTON, self.OnButtonListSave, btn_lst_save)
    
    def OnButtonConvert(self, event):
        """TODO: convert"""
#        print "Convert!"
#        self.Disable()
        convert_ready = False
        a = []
        target_dir = None

        a.append(self.te_size_x.Validate())
        a.append(self.te_size_y.Validate())
        if len(self.plt_list) > 0:
            a.append(True)
        else:
            a.append(False)
        a.append(self.workbook_ok)
        a.append(self.namelist_ok)
        if all(a):
            self.size_x = self.te_size_x.get_result_int()
            self.size_y = self.te_size_y.get_result_int()
            print "size: ", self.size_x, self.size_y
            
            dlg = wx.DirDialog(self, "Choose an output directory:",
                          style=wx.DD_DEFAULT_STYLE
                           #| wx.DD_DIR_MUST_EXIST
                           #| wx.DD_CHANGE_DIR
                           )
            
            if dlg.ShowModal() == wx.ID_OK:
                target_dir = dlg.GetPath()
                convert_ready = True
            else:
                convert_ready = False
            
#        convert_ready = True
        if convert_ready:
            print "converting"
#            proc = commons.SampleThreadProc()
            proc = payload.PayloadThreadProc()
            prg = commons.ThreadProgressDialog(proc, self)
            proc.namelist = self.namelist
            proc.workbook = self.workbook
            proc.source_list = self.plt_list
            proc.save_dir = target_dir
            proc.width    = self.size_x
            proc.height   = self.size_y
            prg.show_and_start()
        else:
            wx.MessageBox('Insufficient data!', 'Error!', wx.OK | wx.ICON_ERROR)
        
#        self.Enable()
    
    def OnButtonChoose(self, event):
        dlg = wx.FileDialog(
            self, message="Select the .xlsx file",
            defaultDir=os.getcwd(), 
            defaultFile="",
            wildcard="Excel 2007 file (*.xlsx)|*.xlsx",
            style=wx.OPEN | wx.CHANGE_DIR
            )
    
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            self.LoadWorkbook(path)
            disp_path = wordwrap(path, self.GetSize()[0], wx.ClientDC(self))
            self.txt_file.SetLabel(disp_path)
        
        dlg.Destroy()
    
    def LoadWorkbook(self, path):
        """membaca file excel
        TODO: loading file di thread lain agar UI nggak macet"""
        self.file_excel    = path
        busy = wx.BusyInfo("Opening {0}".format(path))
        try:
            self.workbook = load_workbook(path)
            self.workbook_ok = True
        except Exception as ex:
            wx.MessageBox('{0}'.format(ex), 'Error!', wx.OK | wx.ICON_ERROR)
            print ex
            self.workbook_ok = False
        busy.Destroy()
    
    def OnButtonChooseNamelist(self, event):
        dlg = wx.FileDialog(
            self, message="Select the namelist.wps file",
            defaultDir=os.getcwd(), 
            defaultFile="",
            wildcard = "namelist.wps file (*.wps)|*.wps|"  \
                      "All files (*.*)|*.*",
            style=wx.OPEN | wx.CHANGE_DIR
            )
    
        if dlg.ShowModal() == wx.ID_OK:
            path                = dlg.GetPath()
            self.file_namelist  = path
            try:
                self.namelist       = namelist_reader.WPSNamelistReader(path)
                self.namelist_ok    = True
            except Exception as ex:
                wx.MessageBox('{0}'.format(ex), 'Error!', wx.OK | wx.ICON_ERROR)
                print ex 
                self.namelist_ok    = False
            
            disp_path = wordwrap(path, self.GetSize()[0], wx.ClientDC(self))
            self.txt_namelist_file.SetLabel(disp_path)
        
        dlg.Destroy()
    
    def OnButtonAddPollutant(self, event):
        dlg = PollutantPicker(self)
        try:
            sh = self.workbook.get_sheet_names()
            dlg.panel.set_sheets(sh)
        except Exception as ex:
            print ex
        
        dlg.CenterOnScreen()
        
        if dlg.ShowModal() == wx.ID_OK:
            pollutant  = dlg.panel.pollutant
            row_start = dlg.panel.row_start
            row_end   = dlg.panel.row_end
            emiss_col = dlg.panel.emission_column
            print pollutant, emiss_col, row_start, ':', emiss_col, row_end
            pollutant_data = dlg.panel.pollutant_data
            
            self.plt_list.append(pollutant_data)
            self.plt_str_list.append(pollutant_data.__str__())
            self.lst_pollutant.SetItems(self.plt_str_list)
    
    def OnButtonEditPollutant(self, event):
        sel = self.lst_pollutant.GetSelection()
        if sel is -1:
            return
        print 'Edit pollutant'
        plt = self.plt_list[sel]
        dlg = PollutantPicker(self)
        try:
            sh = self.workbook.get_sheet_names()
            dlg.panel.set_sheets(sh)
        except Exception as ex:
            print ex
        dlg.panel.set_initial_value(plt)
        dlg.CenterOnScreen()
        
        if dlg.ShowModal() == wx.ID_OK:
            pollutant  = dlg.panel.pollutant
            row_start = dlg.panel.row_start
            row_end   = dlg.panel.row_end
            emiss_col = dlg.panel.emission_column
            print pollutant, emiss_col, row_start, ':', emiss_col, row_end
            pollutant_data = dlg.panel.pollutant_data
            
            self.plt_list[sel] = pollutant_data
            self.plt_str_list[sel] = pollutant_data.__str__()
            self.lst_pollutant.SetItems(self.plt_str_list)
        
        
    
    def OnButtonRemovePollutant(self, event):
        sel = self.lst_pollutant.GetSelection()
        if sel is -1:
            return
        print 'remove! ', sel
        self.plt_list.pop(sel)
        self.plt_str_list.pop(sel)
        self.lst_pollutant.SetItems(self.plt_str_list)
                
    
    def OnButtonAbout(self, event):
        info = wx.AboutDialogInfo()
        info.Name = "WRF Emission Converter"
        info.Version = "0.0.1"
        info.Copyright = "(C) 2011 Arif Widi Nugroho < arif@hexarius.com >"
        info.Description = wordwrap(
            """ Memproses data emisi dari file Excel 2007 ke dalam format WRF/Chem """,
            350, wx.ClientDC(self))
        info.Developers = [ "Arif Widi Nugroho < arif@hexarius.com >"]
        licenseText = "Use it anyway you like!"
        info.License = wordwrap(licenseText, 350, wx.ClientDC(self))

        wx.AboutBox(info)
        
    def OnButtonListOpen(self, event):
        print 'Open list'
        dlg = wx.FileDialog(
            self, message="Open the emission setting file",
            defaultDir=os.getcwd(), 
            defaultFile="",
            wildcard = "Emission Setting File (*.lst)|*.lst|"  \
                      "All files (*.*)|*.*",
            style=wx.OPEN | wx.CHANGE_DIR
            )
    
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            print path
            busy = wx.BusyInfo("Opening {0}".format(path))
            
            try:
                f = open(path, 'r')
                list_data = cPickle.load(f)
                self.plt_list = list_data.plt_list
                self.plt_str_list = list_data.plt_str_list
                self.lst_pollutant.SetItems(self.plt_str_list)
                f.close()
            except Exception as ex:
                wx.MessageBox('{0}'.format(ex), 'Error!', wx.OK | wx.ICON_ERROR)
                print ex
            
            busy.Destroy()
        dlg.Destroy()
    
    def OnButtonListSave(self, event):
        print 'Save list'
        dlg = wx.FileDialog(
            self, message="Save the emission setting file",
            defaultDir=os.getcwd(), 
            defaultFile="",
            wildcard = "Emission Setting File (*.lst)|*.lst|"  \
                      "All files (*.*)|*.*",
            style=wx.SAVE | wx.CHANGE_DIR
            )
    
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            print path
            list_data = PollutantSetting()
            list_data.plt_list = self.plt_list
            list_data.plt_str_list = self.plt_str_list
            busy = wx.BusyInfo("Saving {0}".format(path))
            try:
                f = open(path, 'w')
                cPickle.dump(list_data, f)
                f.close()
            except Exception as ex:
                wx.MessageBox('{0}'.format(ex), 'Error!', wx.OK | wx.ICON_ERROR)
                print ex
            busy.Destroy()
        dlg.Destroy()

class PollutantPicker(wx.Dialog):
    def __init__(self, parent=None, id=-1):
        wx.Dialog.__init__(self, parent, id, "Select Pollutant")
        self.panel = PollutantPickerPanel(self)
        sizer = wx.BoxSizer()
        self.SetSizer(sizer)
        sizer.Add(self.panel, 1, wx.ALL | wx.EXPAND)
        self.Fit()
        
class PollutantPickerPanel(wx.Panel):
    def __init__(self, parent, id=-1):
        wx.Panel.__init__(self, parent, id)
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        self.SetSizer(main_sizer)
        
        btn_ok     = wx.Button(self, wx.ID_OK)
        btn_cancel = wx.Button(self, wx.ID_CANCEL)
        btn_ok.SetDefault()
        sizer_btn = wx.BoxSizer(wx.HORIZONTAL)
        sizer_btn.Add(btn_ok, 0, wx.ALL | wx.EXPAND, 5)
        sizer_btn.Add(wx.BoxSizer(), 1, wx.ALL | wx.EXPAND, 5)
        sizer_btn.Add(btn_cancel, 0, wx.ALL | wx.EXPAND, 5)
        
        plt_list = ['Select Pollutant', 
                    'E_ALD', 'E_CO', 'E_CSL', 'E_ECI', 'E_ECJ', 
                    'E_ETH', 'E_HC3', 'E_HC5', 'E_HC8', 'E_HCHO', 
                    'E_ISO', 'E_KET', 'E_NH3', 'E_NO', 'E_NO3I', 
                    'E_NO3J', 'E_OL2', 'E_OLI', 'E_OLT', 'E_ORA2', 
                    'E_ORGI', 'E_ORGJ', 'E_PM25I', 'E_PM25J', 'E_PM_10', 
                    'E_SO2', 'E_SO4I', 'E_SO4J', 'E_TOL', 'E_XYL']
        self.plt_list = plt_list
        self.choice_plt = wx.ComboBox(self, -1, choices = plt_list)
        self.choice_plt.SetSelection(0)
        txt_label = wx.StaticText(self, -1, "Pollutant: ")
        sizer_plt = wx.BoxSizer(wx.HORIZONTAL)
        sizer_plt.Add(txt_label, 1, wx.ALL | wx.EXPAND, 5)
        sizer_plt.Add(self.choice_plt, 2, wx.ALL | wx.EXPAND, 5)
        
        self.sheet_list = []
        self.choice_sheet = wx.Choice(self, -1, choices = self.sheet_list)
        txt_sheet_label = wx.StaticText(self, -1, "Worksheet: ")
        sizer_sheet = wx.BoxSizer(wx.HORIZONTAL)
        sizer_sheet.Add(txt_sheet_label, 1, wx.ALL | wx.EXPAND, 5)
        sizer_sheet.Add(self.choice_sheet, 2, wx.ALL | wx.EXPAND, 5)
        
        self.te_conv_f      = commons.GrayTextCtrl(self, -1, 'Default value: 1.0)', validator=commons.EvalNumberTextValidator(True))
        sizer_conv_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_conv_header.Add(wx.StaticText(self, -1, "Conversion Factor "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_conv_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        
        self.te_row_start = commons.GrayTextCtrl(self, -1, 'Starting Row (e.g.: 2)', validator=commons.NumberTextValidator())
        self.te_row_end = commons.GrayTextCtrl(self, -1, 'Ending Row (e.g.: 101)', validator=commons.NumberTextValidator())
        sizer_sz = wx.BoxSizer(wx.HORIZONTAL)
        sizer_sz.Add(self.te_row_start, 1, wx.ALL | wx.EXPAND, 5)
        sizer_sz.Add(self.te_row_end, 1, wx.ALL | wx.EXPAND, 5)
        sizer_sz_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_sz_header.Add(wx.StaticText(self, -1, "Data Rows "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_sz_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        
        self.te_emiss_column = commons.GrayTextCtrl(self, -1, 'Column (e.g. A)', validator=commons.AlphaTextValidator())
        sizer_emiss_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_emiss_header.Add(wx.StaticText(self, -1, "Emission Column "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_emiss_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        
        self.te_lat_column = commons.GrayTextCtrl(self, -1, 'Column (e.g. B)', validator=commons.AlphaTextValidator())
        sizer_lat_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_lat_header.Add(wx.StaticText(self, -1, "Latitude Column "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_lat_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        self.te_lon_column = commons.GrayTextCtrl(self, -1, 'Column (e.g. C)', validator=commons.AlphaTextValidator())
        sizer_lon_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_lon_header.Add(wx.StaticText(self, -1, "Longitude Column "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_lon_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        
        self.te_x_column = commons.GrayTextCtrl(self, -1, 'Column (e.g. D)', validator=commons.AlphaTextValidator())
        sizer_x_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_x_header.Add(wx.StaticText(self, -1, "X Column "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_x_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        self.te_y_column = commons.GrayTextCtrl(self, -1, 'Column (e.g. E)', validator=commons.AlphaTextValidator())
        sizer_y_header = wx.BoxSizer(wx.HORIZONTAL)
        sizer_y_header.Add(wx.StaticText(self, -1, "Y Column "), 0, wx.ALL | wx.EXPAND, 0)
        sizer_y_header.Add(wx.StaticLine(self, -1), 1, wx.CENTER, 5)
        
        main_sizer.Add(sizer_plt, 0, wx.ALL | wx.EXPAND, 0)
        main_sizer.Add(sizer_sheet, 0, wx.ALL | wx.EXPAND, 0)
        txt_desc = wordwrap(u"Use conversion factor to convert unit to mol/km\u00b2/hour", 300, wx.ClientDC(self))
        main_sizer.Add(wx.StaticText(self, -1, txt_desc), 0, wx.ALL | wx.EXPAND, 5)
        
        main_sizer.Add(sizer_conv_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(self.te_conv_f, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_sz_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_sz, 0, wx.ALL | wx.EXPAND, 0)
        main_sizer.Add(sizer_emiss_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(self.te_emiss_column, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_lat_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(self.te_lat_column, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_lon_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(self.te_lon_column, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_x_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(self.te_x_column, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_y_header, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(self.te_y_column, 0, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(wx.StaticLine(self, -1), 1, wx.ALL | wx.EXPAND, 5)
        main_sizer.Add(sizer_btn, 0, wx.ALL | wx.EXPAND, 0)
        
        
        self.Bind(wx.EVT_BUTTON, self.OnOk, btn_ok)
    
    def set_sheets(self, sheet_list):
        self.sheet_list = sheet_list
        self.choice_sheet.SetItems(self.sheet_list)
        if len(sheet_list) > 0:
            self.choice_sheet.SetSelection(0)
    
    def set_initial_value(self, pollutant_data):
        sl = self.plt_list.index(pollutant_data.pollutant)
        self.choice_plt.SetSelection(sl)
        self.choice_sheet.SetSelection(pollutant_data.worksheet)
        self.te_conv_f.SetValue(pollutant_data.conversion_factor_str)
        self.te_row_start.SetValue(str(pollutant_data.row_start))
        self.te_row_end.SetValue(str(pollutant_data.row_end))
        self.te_emiss_column.SetValue(str(pollutant_data.emission_column))
        self.te_lat_column.SetValue(str(pollutant_data.lat_column))
        self.te_lon_column.SetValue(str(pollutant_data.lon_column))
        self.te_x_column.SetValue(str(pollutant_data.x_column))
        self.te_y_column.SetValue(str(pollutant_data.y_column))
        
    def OnOk(self, event):
        a = []
        a.append(self.te_conv_f.Validate())
        a.append(self.te_emiss_column.Validate())
        a.append(self.choice_plt.GetStringSelection().find('Select Pollutant') == -1)
        a.append(self.te_row_start.Validate())
        a.append(self.te_row_end.Validate())
        a.append(self.te_lat_column.Validate())
        a.append(self.te_lon_column.Validate())
        a.append(self.te_x_column.Validate())
        a.append(self.te_y_column.Validate())
        self.pollutant         = self.choice_plt.GetStringSelection()
        self.worksheet         = self.choice_sheet.GetSelection()
        self.row_start         = self.te_row_start.get_result_int()
        self.row_end           = self.te_row_end.get_result_int()
        self.emission_column   = self.te_emiss_column.get_result().upper()
        self.lat_column        = self.te_lat_column.get_result().upper()
        self.lon_column        = self.te_lon_column.get_result().upper()
        self.x_column          = self.te_x_column.get_result().upper()
        self.y_column          = self.te_y_column.get_result().upper()
#        self.conversion_factor = self.te_conv_f.get_result_float()
        self.conversion_factor = self.te_conv_f.get_eval_result_float()
        self.conversion_factor_str = self.te_conv_f.get_result()
        if self.conversion_factor == None:
            self.conversion_factor = 1.0
        self.pollutant_data    = PollutantData()
        self.pollutant_data.conversion_factor     = self.conversion_factor
        self.pollutant_data.conversion_factor_str = self.conversion_factor_str
        self.pollutant_data.pollutant             = self.pollutant
        self.pollutant_data.worksheet             = self.worksheet
        self.pollutant_data.row_start             = self.row_start
        self.pollutant_data.row_end               = self.row_end
        self.pollutant_data.lat_column            = self.lat_column
        self.pollutant_data.lon_column            = self.lon_column
        self.pollutant_data.x_column              = self.x_column
        self.pollutant_data.y_column              = self.y_column
        self.pollutant_data.emission_column       = self.emission_column
        if all(a):
            event.Skip()
        return
        

if __name__ == '__main__':
    app = wx.PySimpleApp()
    fr = MainFrame()
    fr.CenterOnScreen()
    fr.Show()
    app.MainLoop()
    
    
